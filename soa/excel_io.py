"""Excel helpers for batch payment report (upload + download)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

_LOAN_HEADER_ALIASES = (
    "loanno",
    "loan_no",
    "loannumber",
    "loan",
    "agreement",
    "agreementno",
    "agreement_no",
    "agreementnumber",
    "lan",
)


def _normalize_header(cell: object) -> str:
    if cell is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(cell).strip().lower())


def find_loan_column_index(header_row: tuple[Any, ...]) -> int | None:
    """0-based column index for loan / agreement number."""
    norm = [_normalize_header(h) for h in header_row]
    for i, h in enumerate(norm):
        if not h:
            continue
        for alias in _LOAN_HEADER_ALIASES:
            if h == alias or alias in h or h in alias:
                return i
    return None


def read_loan_numbers_from_xlsx(content: bytes) -> list[str]:
    """Read first sheet; detect loan column by header name; return unique uppercase IDs."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return []
        idx = find_loan_column_index(tuple(header))
        if idx is None:
            raise ValueError(
                "Could not find a loan column. Use a header like 'loan no' or 'agreement'."
            )
        seen: set[str] = set()
        out: list[str] = []
        for row in rows:
            if not row or idx >= len(row):
                continue
            raw = row[idx]
            if raw is None:
                continue
            s = str(raw).strip().upper()
            s = re.sub(r"[^A-Z0-9]", "", s)
            if len(s) < 5 or len(s) > 40 or s in seen:
                continue
            seen.add(s)
            out.append(s)
        return out
    finally:
        wb.close()


def write_payment_report_xlsx(
    payment_rows: list[dict[str, Any]],
    loan_status_rows: list[dict[str, Any]],
) -> bytes:
    """Build workbook: sheet Payments + sheet LoanStatus."""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Payments"
    headers = (
        "loan_no",
        "date",
        "amount",
        "payment_type",
        "reference",
        "source",
        "particulars",
    )
    ws0.append(list(headers))
    for r in payment_rows:
        ws0.append(
            [
                r.get("loan_no"),
                r.get("date"),
                r.get("amount"),
                r.get("payment_type"),
                r.get("reference", ""),
                r.get("source"),
                (r.get("particulars") or "")[:500],
            ]
        )

    ws1 = wb.create_sheet("Loan status")
    ws1.append(["loan_no", "status", "detail", "payments_found"])
    for r in loan_status_rows:
        ws1.append(
            [
                r.get("loan_no"),
                r.get("status"),
                r.get("detail", ""),
                r.get("payments_found", ""),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
